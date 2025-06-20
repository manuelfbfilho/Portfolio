from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font

# carregando o arquivo Excel
wb = load_workbook("Modificando estrutura.xlsx")

# Verificando as abas
print(wb.sheetnames)

# selecionando a aba Vendas
sheet = wb["Vendas"]

# renomeando uma aba
sheet.title = "Vendas 2024"

# criando uma nova aba
wb.create_sheet("Vendas 2025")

# mesclando células
sheet.merge_cells("A1:D1")

# Centralizar e deixar o texto em negrito
sheet["A1"].alignment, sheet["A1"].font = Alignment(horizontal="center", vertical="center"), Font(bold=True)

# tirando a mesclagem de células
sheet.unmerge_cells("A1:D1")

# inserindo linhas
sheet.insert_rows(3)
sheet.insert_rows(14)

# inserindo colunas
sheet.insert_cols(2)

# deletando linhas e colunas
sheet.delete_rows(3)
sheet.delete_cols(2)

# salvando o arquivo
wb.save("Modificando estrutura.xlsx")



