from openpyxl import load_workbook
from prettytable import PrettyTable

wb = load_workbook("Lendo Excel.xlsx")

# verificando as abas
print(wb.sheetnames)

# acessando as abas
sheet_nomes = wb["Nomes"]
sheet_semanas = wb["Dias da semana"]
sheet_doces = wb["Doces"]

# acessando os dados das células
print(sheet_nomes["A2"].value)
print(sheet_semanas.cell(8, 1).value)

# atualizando o valor de uma célula
sheet_doces["A6"].value = "Bomba doce de leite"

# descobrindo o total de linhas e de colunas
print(sheet_nomes.max_row)
print(sheet_nomes.max_column)

# Crie uma nova tabela
tabela = PrettyTable()

# Defina os nomes das colunas
tabela.field_names = ["Coluna 1", "Coluna 2"]

for linha in range(2, sheet_nomes.max_row + 1):
    # Crie uma lista para armazenar os dados da linha
    dados_linha = []
    for coluna in range(1, sheet_nomes.max_column + 1):
        dados_linha.append(sheet_nomes.cell(linha, coluna).value)
    # Adicione os dados à tabela
    tabela.add_row(dados_linha)

# Imprima a tabela
print(tabela)

# percorrendo os dados de uma aba
for linha in range(2, sheet_nomes.max_row + 1):
    for coluna in range(1, sheet_nomes.max_column + 1):
        print(sheet_nomes.cell(linha, coluna).value)

# salvando o arquivo
wb.save("Lendo Excel.xlsx")

