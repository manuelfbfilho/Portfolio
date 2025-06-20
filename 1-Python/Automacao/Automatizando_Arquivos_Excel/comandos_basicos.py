from openpyxl import Workbook

# criando um Arquivo Excel
wb = Workbook()

# Verificando as abas
print(wb.sheetnames)

# selecionando uma aba
sheet = wb["Sheet"]

# inserindo dados pelo rótulo da célula
sheet["A1"].value = "Python"
sheet["B1"].value = "Formação Expert"
sheet["C1"].value = "Automação de processos"

# inserindo dados pela coordenada da célula
sheet.cell(row=2, column=1).value = "Empowerdata"
sheet.cell(2, 2).value = "Empowerpython"

# apagando o dado de uma célula
sheet["A1"].value = None
sheet.cell(1, 3).value = None

wb.save("Comando básicos.xlsx")