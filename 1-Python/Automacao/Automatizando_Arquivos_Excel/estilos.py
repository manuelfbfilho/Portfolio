from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = load_workbook("Estilos.xlsx")
sheet = wb["Estilos"]

# alterando a cor
sheet["A1"].font = Font(color="000000FF")

# alterando o preenchimento
sheet["A2"].fill = PatternFill("darkDown", "00FFFF00")

# alterando o tipo de fonte
sheet["A3"].font = Font(name="Times New Roman")

# aplicando negrito e itálico
sheet["A4"].font = Font(bold=True)
sheet["A5"].font = Font(italic=True)

# Alterando o alinhamento
sheet["A6"].alignment = Alignment(horizontal="center", vertical="center")

# salvando o arquivo
wb.save("Estilos.xlsx")

